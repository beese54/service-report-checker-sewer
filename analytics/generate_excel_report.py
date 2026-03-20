"""
analytics/generate_excel_report.py
====================================
Reads the output_data dict from run_pipeline() and writes an Excel workbook
with two sheets:

  - "Results"    : one row per folder, colour-coded pass/fail cells,
                   merged group headers, frozen panes
  - "Histograms" : 8 matplotlib figures embedded as PNG images

Usage:
    from analytics.generate_excel_report import generate_excel_report
    generate_excel_report(output_data, "/path/to/pipeline_report.xlsx")
"""

import io
import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Fill colours ──────────────────────────────────────────────────────────────
_F_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_F_ORANGE = PatternFill("solid", fgColor="FFEB9C")
_F_RED    = PatternFill("solid", fgColor="FFC7CE")
_F_GREY   = PatternFill("solid", fgColor="D9D9D9")

_BOLD        = Font(bold=True)
_WHITE_BOLD  = Font(bold=True, color="FFFFFF")
_CENTER      = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Group header background colours (dark; white text)
_GROUP_FILLS = {
    "Identity":        PatternFill("solid", fgColor="1F3864"),
    "Gate -- D":       PatternFill("solid", fgColor="833C00"),
    "Gate -- U":       PatternFill("solid", fgColor="833C00"),
    "Stage 2N -- D":   PatternFill("solid", fgColor="1F497D"),
    "Stage 2N -- U":   PatternFill("solid", fgColor="1F497D"),
    "Stage 3N -- D2":  PatternFill("solid", fgColor="4B2B6B"),
    "Stage 3N -- U2":  PatternFill("solid", fgColor="4B2B6B"),
    "Stage 4N -- D3":  PatternFill("solid", fgColor="0D4E1A"),
    "Stage 4N -- U3":  PatternFill("solid", fgColor="0D4E1A"),
}

# ── Column definitions ─────────────────────────────────────────────────────────
# (display_header, data_key, group_name, col_width)
_COLUMNS = [
    # ── Identity ─────────────────────────────────────────────────────────────
    ("Folder",        "Folder",         "Identity",        22),
    ("Type",          "Type",           "Identity",        14),
    ("Status",        "Status",         "Identity",        20),
    ("Failed Stage",  "Failed Stage",   "Identity",        14),
    # ── Gate D ───────────────────────────────────────────────────────────────
    ("D1 Conf",       "D1 Gate Conf",   "Gate -- D",        9),
    ("D4 Conf",       "D4 Gate Conf",   "Gate -- D",        9),
    ("D Gate Tier",   "D Gate Tier",    "Gate -- D",       11),
    # ── Gate U ───────────────────────────────────────────────────────────────
    ("U1 Conf",       "U1 Gate Conf",   "Gate -- U",        9),
    ("U4 Conf",       "U4 Gate Conf",   "Gate -- U",        9),
    ("U Gate Tier",   "U Gate Tier",    "Gate -- U",       11),
    # ── Stage 2N D ───────────────────────────────────────────────────────────
    ("Status",        "D2N Status",     "Stage 2N -- D",   11),
    ("KP Before",     "D KP Before",    "Stage 2N -- D",    9),
    ("KP After",      "D KP After",     "Stage 2N -- D",    9),
    ("FLANN",         "D FLANN",        "Stage 2N -- D",    8),
    ("Inliers",       "D Inliers",      "Stage 2N -- D",    8),
    ("Cover%",        "D Coverage%",    "Stage 2N -- D",    8),
    ("Pass",          "D S2N Pass",     "Stage 2N -- D",    7),
    # ── Stage 2N U ───────────────────────────────────────────────────────────
    ("Status",        "U2N Status",     "Stage 2N -- U",   11),
    ("KP Before",     "U KP Before",    "Stage 2N -- U",    9),
    ("KP After",      "U KP After",     "Stage 2N -- U",    9),
    ("FLANN",         "U FLANN",        "Stage 2N -- U",    8),
    ("Inliers",       "U Inliers",      "Stage 2N -- U",    8),
    ("Cover%",        "U Coverage%",    "Stage 2N -- U",    8),
    ("Pass",          "U S2N Pass",     "Stage 2N -- U",    7),
    ("S2N Overall",   "S2N Overall",    "Stage 2N -- U",   11),
    # ── Stage 3N D2 ──────────────────────────────────────────────────────────
    ("Status",        "D2 Status",      "Stage 3N -- D2",  10),
    ("Wash Tier",     "D2 Wash Tier",   "Stage 3N -- D2",  10),
    ("Wash Conf",     "D2 Wash Conf",   "Stage 3N -- D2",   9),
    ("KP Ratio",      "D2 KP Ratio",    "Stage 3N -- D2",   8),
    ("Std Incr%",     "D2 Std Incr%",   "Stage 3N -- D2",   9),
    ("Entropy",       "D2 Entropy",     "Stage 3N -- D2",   8),
    ("Edge Incr%",    "D2 Edge Incr%",  "Stage 3N -- D2",   9),
    ("Mean Diff",     "D2 Mean Diff",   "Stage 3N -- D2",   9),
    ("Pct Chgd",      "D2 Pct Chgd",    "Stage 3N -- D2",   9),
    ("Inliers",       "D2 Inliers",     "Stage 3N -- D2",   8),
    # ── Stage 3N U2 ──────────────────────────────────────────────────────────
    ("Status",        "U2 Status",      "Stage 3N -- U2",  10),
    ("Wash Tier",     "U2 Wash Tier",   "Stage 3N -- U2",  10),
    ("Wash Conf",     "U2 Wash Conf",   "Stage 3N -- U2",   9),
    ("KP Ratio",      "U2 KP Ratio",    "Stage 3N -- U2",   8),
    ("Std Incr%",     "U2 Std Incr%",   "Stage 3N -- U2",   9),
    ("Entropy",       "U2 Entropy",     "Stage 3N -- U2",   8),
    ("Edge Incr%",    "U2 Edge Incr%",  "Stage 3N -- U2",   9),
    ("Mean Diff",     "U2 Mean Diff",   "Stage 3N -- U2",   9),
    ("Pct Chgd",      "U2 Pct Chgd",    "Stage 3N -- U2",   9),
    ("Inliers",       "U2 Inliers",     "Stage 3N -- U2",   8),
    ("S3N Overall",   "S3N Overall",    "Stage 3N -- U2",  11),
    # ── Stage 4N D3 ──────────────────────────────────────────────────────────
    ("Status",        "D3 Status",      "Stage 4N -- D3",  10),
    ("Score",         "D3 Score",       "Stage 4N -- D3",   7),
    ("Grease%",       "D3 Grease%",     "Stage 4N -- D3",   9),
    ("Texture OK",    "D3 Texture OK",  "Stage 4N -- D3",   9),
    ("Water",         "D3 Water",       "Stage 4N -- D3",   7),
    ("Water Conf",    "D3 Water Conf",  "Stage 4N -- D3",   9),
    # ── Stage 4N U3 ──────────────────────────────────────────────────────────
    ("Status",        "U3 Status",      "Stage 4N -- U3",  10),
    ("Score",         "U3 Score",       "Stage 4N -- U3",   7),
    ("Grease%",       "U3 Grease%",     "Stage 4N -- U3",   9),
    ("Texture OK",    "U3 Texture OK",  "Stage 4N -- U3",   9),
    ("Water",         "U3 Water",       "Stage 4N -- U3",   7),
    ("Water Conf",    "U3 Water Conf",  "Stage 4N -- U3",   9),
    ("S4N Verdict",   "S4N Verdict",    "Stage 4N -- U3",  12),
]


# ── Row extraction ─────────────────────────────────────────────────────────────

def _extract_row(r):
    """Flatten one folder result dict into a key→value mapping."""
    d  = r.get("detail", {})
    s2 = d.get("stage2n") or {}
    s3 = d.get("stage3n") or {}
    s4 = d.get("stage4n") or {}
    # Obstruction folders may run SIFT stages in a sub-dict
    if "sift_stages" in d:
        ss = d["sift_stages"]
        s2 = ss.get("stage2n") or {}
        s3 = ss.get("stage3n") or {}
        s4 = ss.get("stage4n") or {}

    dp  = (s2.get("pair_stats") or {}).get("D") or {}
    up  = (s2.get("pair_stats") or {}).get("U") or {}
    d2p = (s3.get("pair_results") or {}).get("D2") or {}
    u2p = (s3.get("pair_results") or {}).get("U2") or {}
    d3p = (s4.get("pair_results") or {}).get("D3") or {}
    u3p = (s4.get("pair_results") or {}).get("U3") or {}

    def _pair_gate_tier(p, b_key, a_key):
        """Derive worst gate tier for a pair from status or per-image keys."""
        s = p.get("status") or ""
        if s.startswith("GATE_"):
            return s[5:]          # "REVIEW" or "REJECT"
        t1 = (p.get(f"gate_tier_{b_key}") or "").lower()
        t2 = (p.get(f"gate_tier_{a_key}") or "").lower()
        for t in (t1, t2):
            if t == "reject":
                return "REJECT"
        for t in (t1, t2):
            if t == "review":
                return "REVIEW"
        if s == "OK" and (t1 or t2):
            return "ACCEPT"
        return None

    return {
        # Identity
        "Folder":         r.get("name"),
        "Type":           r.get("type"),
        "Status":         r.get("status"),
        "Failed Stage":   r.get("failed_stage"),
        # Gate D
        "D1 Gate Conf":   dp.get("gate_conf_D1"),
        "D4 Gate Conf":   dp.get("gate_conf_D4"),
        "D Gate Tier":    _pair_gate_tier(dp, "D1", "D4"),
        # Gate U
        "U1 Gate Conf":   up.get("gate_conf_U1"),
        "U4 Gate Conf":   up.get("gate_conf_U4"),
        "U Gate Tier":    _pair_gate_tier(up, "U1", "U4"),
        # Stage 2N D
        "D2N Status":     dp.get("status"),
        "D KP Before":    dp.get("kp_before"),
        "D KP After":     dp.get("kp_after"),
        "D FLANN":        dp.get("ratio_matches"),
        "D Inliers":      dp.get("ransac_inliers"),
        "D Coverage%":    dp.get("inlier_coverage_pct"),
        "D S2N Pass":     dp.get("stage2n_pass"),
        # Stage 2N U
        "U2N Status":     up.get("status"),
        "U KP Before":    up.get("kp_before"),
        "U KP After":     up.get("kp_after"),
        "U FLANN":        up.get("ratio_matches"),
        "U Inliers":      up.get("ransac_inliers"),
        "U Coverage%":    up.get("inlier_coverage_pct"),
        "U S2N Pass":     up.get("stage2n_pass"),
        "S2N Overall":    s2.get("overall_pass"),
        # Stage 3N D2
        "D2 Status":      d2p.get("status"),
        "D2 Wash Tier":   d2p.get("washing_tier"),
        "D2 Wash Conf":   d2p.get("washing_confidence"),
        "D2 KP Ratio":    (d2p.get("metrics") or {}).get("kp_ratio"),
        "D2 Std Incr%":   (d2p.get("metrics") or {}).get("std_increase_pct"),
        "D2 Entropy":     (d2p.get("metrics") or {}).get("entropy_increase"),
        "D2 Edge Incr%":  (d2p.get("metrics") or {}).get("edge_increase_pct"),
        "D2 Mean Diff":   (d2p.get("sift_stats") or {}).get("mean_diff"),
        "D2 Pct Chgd":    (d2p.get("sift_stats") or {}).get("pct_changed"),
        "D2 Inliers":     (d2p.get("sift_stats") or {}).get("ransac_inliers"),
        # Stage 3N U2
        "U2 Status":      u2p.get("status"),
        "U2 Wash Tier":   u2p.get("washing_tier"),
        "U2 Wash Conf":   u2p.get("washing_confidence"),
        "U2 KP Ratio":    (u2p.get("metrics") or {}).get("kp_ratio"),
        "U2 Std Incr%":   (u2p.get("metrics") or {}).get("std_increase_pct"),
        "U2 Entropy":     (u2p.get("metrics") or {}).get("entropy_increase"),
        "U2 Edge Incr%":  (u2p.get("metrics") or {}).get("edge_increase_pct"),
        "U2 Mean Diff":   (u2p.get("sift_stats") or {}).get("mean_diff"),
        "U2 Pct Chgd":    (u2p.get("sift_stats") or {}).get("pct_changed"),
        "U2 Inliers":     (u2p.get("sift_stats") or {}).get("ransac_inliers"),
        "S3N Overall":    s3.get("overall_pass"),
        # Stage 4N D3
        "D3 Status":      d3p.get("status"),
        "D3 Score":       d3p.get("score"),
        "D3 Grease%":     (d3p.get("grease") or {}).get("grease_pct"),
        "D3 Texture OK":  (d3p.get("texture") or {}).get("confirmed"),
        "D3 Water":       (d3p.get("water") or {}).get("water_detected"),
        "D3 Water Conf":  (d3p.get("water") or {}).get("water_confidence"),
        # Stage 4N U3
        "U3 Status":      u3p.get("status"),
        "U3 Score":       u3p.get("score"),
        "U3 Grease%":     (u3p.get("grease") or {}).get("grease_pct"),
        "U3 Texture OK":  (u3p.get("texture") or {}).get("confirmed"),
        "U3 Water":       (u3p.get("water") or {}).get("water_detected"),
        "U3 Water Conf":  (u3p.get("water") or {}).get("water_confidence"),
        "S4N Verdict":    s4.get("folder_verdict"),
    }


# ── Cell colour logic ──────────────────────────────────────────────────────────

def _cell_fill(value, data_key):
    """Return a PatternFill for a data cell, or None for default."""
    if value is None:
        return _F_GREY

    k = data_key

    if k == "Status":
        s = str(value).upper()
        if s == "ACCEPTED":               return _F_GREEN
        if s == "NEEDS_REVIEW":           return _F_ORANGE
        if s in ("REJECTED", "OBSTRUCTION_PROCESSED"):
                                          return _F_RED
        return None

    if k in ("S4N Verdict",):
        s = str(value).upper()
        if s == "ACCEPTED":   return _F_GREEN
        if s == "NEEDS_REVIEW": return _F_ORANGE
        if s == "REJECTED":   return _F_RED
        return None

    if "Gate Tier" in k:
        s = str(value).upper()
        if s in ("ACCEPT",):              return _F_GREEN
        if s in ("REVIEW",):              return _F_ORANGE
        if s in ("REJECT",):              return _F_RED
        return None

    if "Wash Tier" in k:
        s = str(value).upper()
        if s == "HIGH":    return _F_GREEN
        if s == "MEDIUM":  return _F_ORANGE
        if s == "LOW":     return _F_RED
        return None

    if k in ("D3 Status", "U3 Status"):
        s = str(value).upper()
        if s == "PASS":    return _F_GREEN
        if s == "REVIEW":  return _F_ORANGE
        if s in ("FAIL",): return _F_RED
        return None

    if k in ("D2N Status", "U2N Status", "D2 Status", "U2 Status"):
        s = str(value).upper()
        if s == "OK":      return _F_GREEN
        if s == "FAILED":  return _F_RED
        return None

    if k in (
        "D S2N Pass", "U S2N Pass", "S2N Overall", "S3N Overall",
        "D3 Texture OK", "U3 Texture OK",
    ):
        if value is True:  return _F_GREEN
        if value is False: return _F_RED
        return None

    if k in ("D3 Water", "U3 Water"):
        if value is True:  return _F_GREEN
        if value is False: return _F_RED
        return None

    return None


def _row_fill(status):
    """Light row-level highlight based on folder status."""
    if status == "ACCEPTED":
        return PatternFill("solid", fgColor="EBF1DE")
    if status == "REJECTED":
        return PatternFill("solid", fgColor="FFE4E4")
    return None


# ── Results sheet ──────────────────────────────────────────────────────────────

def _build_results_sheet(ws, rows):
    """Write the Results sheet with merged group headers and colour coding."""
    headers   = [c[0] for c in _COLUMNS]
    data_keys = [c[1] for c in _COLUMNS]
    groups    = [c[2] for c in _COLUMNS]
    widths    = [c[3] for c in _COLUMNS]

    # ── Row 1: group header (merged cells) ────────────────────────────────────
    col = 1
    i   = 0
    while i < len(groups):
        g     = groups[i]
        start = i + 1
        end   = i + 1
        while end < len(groups) and groups[end] == g:
            end += 1
        end_col = end  # exclusive

        start_letter = get_column_letter(start)
        end_letter   = get_column_letter(end_col)

        if start == end_col:
            cell = ws.cell(row=1, column=start, value=g)
        else:
            ws.merge_cells(f"{start_letter}1:{end_letter}1")
            cell = ws.cell(row=1, column=start, value=g)

        cell.fill      = _GROUP_FILLS.get(g, PatternFill("solid", fgColor="333333"))
        cell.font      = _WHITE_BOLD
        cell.alignment = _CENTER

        i = end_col

    # ── Row 2: column headers ─────────────────────────────────────────────────
    for col_idx, (hdr, width) in enumerate(zip(headers, widths), start=1):
        cell            = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font       = _BOLD
        cell.alignment  = _CENTER_WRAP
        cell.fill       = PatternFill("solid", fgColor="BDD7EE")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30

    # ── Data rows (starting row 3) ─────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=3):
        status   = row_data.get("Status")
        row_fill = _row_fill(status)

        for col_idx, key in enumerate(data_keys, start=1):
            value = row_data.get(key)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)

            # Per-cell colour overrides row highlight
            fill = _cell_fill(value, key)
            if fill is not None:
                cell.fill = fill
            elif row_fill is not None:
                cell.fill = row_fill

            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Freeze top-left (row 2 header + column A) ─────────────────────────────
    ws.freeze_panes = "B3"

    # ── Auto-filter on row 2 ──────────────────────────────────────────────────
    last_col = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A2:{last_col}2"


# ── Histogram helpers ──────────────────────────────────────────────────────────

def _fig_to_xl_image(fig, dpi=100):
    """Convert a matplotlib Figure to an openpyxl Image object via BytesIO."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return XLImage(buf)


def _vals(rows, key):
    """Extract non-None numeric values for a data key."""
    return [v for v in (r.get(key) for r in rows) if v is not None]


def _build_histograms_sheet(ws, rows):
    """Embed 8 matplotlib histogram figures into the Histograms sheet."""
    plt.style.use("dark_background")

    row_anchor = 1   # current top row for next image block

    def _embed(fig, col_letter="A"):
        nonlocal row_anchor
        img = _fig_to_xl_image(fig)
        ws.add_image(img, f"{col_letter}{row_anchor}")
        # Approximate row advance: figure height in pts / 15 pts-per-row + 2 gap
        h_px = fig.get_size_inches()[1] * 100
        row_advance = max(int(h_px / 15) + 2, 25)
        row_anchor += row_advance

    # ── 1. Gate confidence — 4-panel ─────────────────────────────────────────
    fig, axes = plt.subplots(1, 4, figsize=(16, 4), tight_layout=True)
    fig.suptitle("Gate Confidence", fontsize=13)
    for ax, key, label in zip(
        axes,
        ["D1 Gate Conf", "D4 Gate Conf", "U1 Gate Conf", "U4 Gate Conf"],
        ["D1", "D4", "U1", "U4"],
    ):
        data = _vals(rows, key)
        if data:
            ax.hist(data, bins=20, range=(0, 1), color="#4FC3F7", edgecolor="none")
        ax.axvline(0.30, color="#FFEB9C", linewidth=1.5, linestyle="--", label="Review 0.30")
        ax.axvline(0.70, color="#C6EFCE", linewidth=1.5, linestyle="--", label="Accept 0.70")
        ax.set_title(label)
        ax.set_xlabel("Confidence")
        ax.set_ylabel("Count")
        ax.legend(fontsize=7)
    _embed(fig)

    # ── 2. RANSAC Inliers — 2-panel ───────────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), tight_layout=True)
    fig.suptitle("RANSAC Inliers (Stage 2N)", fontsize=13)
    for ax, key, label in zip(axes, ["D Inliers", "U Inliers"], ["D pair", "U pair"]):
        data = _vals(rows, key)
        if data:
            ax.hist(data, bins=30, color="#80CBC4", edgecolor="none")
        ax.axvline(10, color="#FFC7CE", linewidth=1.5, linestyle="--", label="Threshold 10")
        ax.set_title(label)
        ax.set_xlabel("Inliers")
        ax.set_ylabel("Count")
        ax.legend(fontsize=8)
    _embed(fig)

    # ── 3. Inlier Coverage % — 2-panel ────────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), tight_layout=True)
    fig.suptitle("Inlier Coverage % (Stage 2N)", fontsize=13)
    for ax, key, label in zip(axes, ["D Coverage%", "U Coverage%"], ["D pair", "U pair"]):
        data = _vals(rows, key)
        if data:
            ax.hist(data, bins=25, range=(0, 100), color="#CE93D8", edgecolor="none")
        ax.axvline(25, color="#FFC7CE", linewidth=1.5, linestyle="--", label="Threshold 25%")
        ax.set_title(label)
        ax.set_xlabel("Coverage %")
        ax.set_ylabel("Count")
        ax.legend(fontsize=8)
    _embed(fig)

    # ── 4. Washing Confidence — 2-panel ───────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), tight_layout=True)
    fig.suptitle("Washing Confidence (Stage 3N)", fontsize=13)
    for ax, key, label in zip(
        axes, ["D2 Wash Conf", "U2 Wash Conf"], ["D2", "U2"]
    ):
        data = _vals(rows, key)
        if data:
            ax.hist(data, bins=20, range=(0, 1), color="#FFB74D", edgecolor="none")
        ax.axvline(0.50, color="#C6EFCE", linewidth=1.5, linestyle="--", label="HIGH 0.50")
        ax.axvline(0.30, color="#FFEB9C", linewidth=1.5, linestyle="--", label="MEDIUM 0.30")
        ax.set_title(label)
        ax.set_xlabel("Confidence")
        ax.set_ylabel("Count")
        ax.legend(fontsize=8)
    _embed(fig)

    # ── 5. Washing Tier Counts — 2 grouped bar charts ─────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), tight_layout=True)
    fig.suptitle("Washing Tier Counts (Stage 3N)", fontsize=13)
    tiers  = ["HIGH", "MEDIUM", "LOW"]
    colors = ["#C6EFCE", "#FFEB9C", "#FFC7CE"]
    for ax, key, label in zip(
        axes, ["D2 Wash Tier", "U2 Wash Tier"], ["D2", "U2"]
    ):
        tier_vals = [str(v).upper() for v in (r.get(key) for r in rows) if v is not None]
        counts    = [tier_vals.count(t) for t in tiers]
        bars      = ax.bar(tiers, counts, color=colors, edgecolor="none")
        for bar, cnt in zip(bars, counts):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.3,
                str(cnt), ha="center", va="bottom", fontsize=9,
            )
        ax.set_title(label)
        ax.set_ylabel("Count")
    _embed(fig)

    # ── 6. Stage 4N Score — 2-panel ───────────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), tight_layout=True)
    fig.suptitle("Stage 4N Score (0-5)", fontsize=13)
    for ax, key, label in zip(axes, ["D3 Score", "U3 Score"], ["D3", "U3"]):
        data = _vals(rows, key)
        if data:
            ax.hist(data, bins=range(7), color="#A5D6A7", edgecolor="#1b5e20",
                    align="left", rwidth=0.7)
        ax.axvline(2.5, color="#FFC7CE", linewidth=1.5, linestyle="--",
                   label="Pass >= 3")
        ax.set_title(label)
        ax.set_xlabel("Score")
        ax.set_ylabel("Count")
        ax.set_xticks(range(6))
        ax.legend(fontsize=8)
    _embed(fig)

    # ── 7. Stage Pass Rates — horizontal bar ──────────────────────────────────
    n = len(rows)
    stages = [
        ("S2N Overall",  "Stage 2N"),
        ("S3N Overall",  "Stage 3N"),
        ("D3 Status",    "Stage 4N D"),
        ("U3 Status",    "Stage 4N U"),
    ]
    labels = [s[1] for s in stages]
    pass_rates = []
    for key, _ in stages:
        vals = [r.get(key) for r in rows]
        if key in ("D3 Status", "U3 Status"):
            passed = sum(1 for v in vals if str(v or "").upper() == "PASS")
        else:
            passed = sum(1 for v in vals if v is True)
        pass_rates.append(passed / n * 100 if n else 0)

    fig, ax = plt.subplots(figsize=(9, 4), tight_layout=True)
    fig.suptitle("Stage Pass Rates (%)", fontsize=13)
    bar_colors = [
        "#C6EFCE" if p >= 70 else "#FFEB9C" if p >= 40 else "#FFC7CE"
        for p in pass_rates
    ]
    bars = ax.barh(labels, pass_rates, color=bar_colors, edgecolor="none")
    for bar, rate in zip(bars, pass_rates):
        ax.text(
            bar.get_width() + 1, bar.get_y() + bar.get_height() / 2,
            f"{rate:.1f}%", va="center", fontsize=9,
        )
    ax.set_xlim(0, 110)
    ax.set_xlabel("% of all folders")
    _embed(fig)

    # ── 8. Final Status Distribution ──────────────────────────────────────────
    status_order  = ["ACCEPTED", "NEEDS_REVIEW", "REJECTED", "OBSTRUCTION_PROCESSED"]
    status_colors = ["#C6EFCE",  "#FFEB9C",       "#FFC7CE",  "#BDD7EE"]
    status_vals   = [str(r.get("Status") or "").upper() for r in rows]
    counts        = [status_vals.count(s) for s in status_order]

    fig, ax = plt.subplots(figsize=(9, 4), tight_layout=True)
    fig.suptitle("Final Status Distribution", fontsize=13)
    bars = ax.bar(status_order, counts, color=status_colors, edgecolor="none")
    for bar, cnt in zip(bars, counts):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.3,
            str(cnt), ha="center", va="bottom", fontsize=10,
        )
    ax.set_ylabel("Folders")
    ax.set_xticklabels(status_order, rotation=15, ha="right")
    _embed(fig)

    # Reset style
    plt.style.use("default")


# ── Public entry point ─────────────────────────────────────────────────────────

def generate_excel_report(output_data, excel_path):
    """
    Generate an Excel workbook from a run_pipeline() output_data dict.

    Parameters
    ----------
    output_data : dict  The dict returned by run_pipeline() (or loaded from JSON).
    excel_path  : str   Full path for the .xlsx file to write.
    """
    folders = output_data.get("folders", [])
    rows    = [_extract_row(r) for r in folders]

    wb = Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────────────────────
    ws_results = wb.active
    ws_results.title = "Results"
    _build_results_sheet(ws_results, rows)

    # ── Sheet 2: Histograms ───────────────────────────────────────────────────
    ws_histo = wb.create_sheet("Histograms")
    ws_histo.column_dimensions["A"].width = 140
    _build_histograms_sheet(ws_histo, rows)

    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)
    wb.save(excel_path)
    return excel_path
